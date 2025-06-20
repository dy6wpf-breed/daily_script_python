import sys
import time
from Bio import Entrez
from openpyxl import load_workbook, Workbook
import os
import xml.etree.ElementTree as ET

Entrez.email = "bjlmwpf@163.com"  # 请替换为您的邮箱地址

def get_snp_info(gene_id):
    try:
        print(f"正在从NCBI获取Gene ID {gene_id}的SNP信息...")
        # 首先获取与该基因相关的SNP ID
        handle = Entrez.elink(dbfrom="gene", db="snp", id=gene_id, linkname="gene_snp", organism="Bos taurus")
        record = Entrez.read(handle)
        handle.close()

        if not record or not record[0].get('LinkSetDb'):
            print(f"Gene ID {gene_id} 没有关联的SNP信息。")
            return None

        snp_ids = [link['Id'] for link in record[0]['LinkSetDb'][0]['Link']]

        if not snp_ids:
            print(f"Gene ID {gene_id} 没有找到相关的SNP ID。")
            return None

        print(f"找到 {len(snp_ids)} 个相关的SNP。")

        # 获取SNP详细信息
        handle = Entrez.efetch(db="snp", id=",".join(snp_ids[:5]), retmode="xml")  # 限制为前5个SNP以加快处理
        xml_data = handle.read()
        handle.close()
        return xml_data
    except Exception as e:
        print(f"获取Gene ID {gene_id}的SNP信息时出错：{str(e)}")
        return None

def extract_snp_info(xml_data):
    root = ET.fromstring(xml_data)
    snps = []

    for rs in root.findall(".//Rs"):
        snp_info = {
            "rs_id": rs.get("rsId"),
            "chromosome": "",
            "position": "",
            "alleles": "",
            "gene_name": "",
            "clinical_significance": "",
            "phenotypes": []
        }

        # 提取染色体和位置信息
        assembly = rs.find(".//Assembly[@genomeBuild='Bos_taurus_ARS-UCD1.2']")
        if assembly is not None:
            component = assembly.find(".//Component")
            if component is not None:
                snp_info["chromosome"] = component.get("chromosome")
                snp_info["position"] = component.get("position")

        # 提取等位基因信息
        observed = rs.find(".//Observed")
        if observed is not None:
            snp_info["alleles"] = observed.text

        # 提取基因名称
        gene_name = rs.find(".//Gene/NAME")
        if gene_name is not None:
            snp_info["gene_name"] = gene_name.text

        # 提取临床意义
        clinical_sig = rs.find(".//ClinicalSignificance")
        if clinical_sig is not None:
            snp_info["clinical_significance"] = clinical_sig.text

        # 提取关联表型
        phenotypes = rs.findall(".//PhenotypeList/Phenotype")
        for phenotype in phenotypes:
            snp_info["phenotypes"].append(phenotype.get("name"))

        snps.append(snp_info)

    return snps

def main():
    try:
        print("程序开始执行...")

        input_excel_file = r"E:\desk\基因注释1.xlsx"
        output_excel_file = r"E:\desk\牛SNP信息结果.xlsx"
        print(f"输入Excel文件路径: {input_excel_file}")
        print(f"输出Excel文件路径: {output_excel_file}")

        if not os.path.exists(input_excel_file):
            print(f"错误：找不到文件 '{input_excel_file}'")
            return

        print(f"正在打开输入Excel文件: {input_excel_file}")
        input_workbook = load_workbook(filename=input_excel_file)
        input_sheet = input_workbook.active

        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(["Gene ID", "RS ID", "染色体", "位置", "等位基因", "基因名称", "临床意义", "关联表型"])

        print(f"Excel文件打开成功。开始读取数据...")
        for row in input_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            gene_id = str(row[0])
            print(f"\n正在处理Gene ID: {gene_id}")

            xml_data = get_snp_info(gene_id)
            if xml_data:
                try:
                    snps = extract_snp_info(xml_data)
                    if not snps:
                        print(f"Gene ID {gene_id} 没有提取到SNP信息。")
                        continue

                    for snp in snps:
                        output_sheet.append([
                            gene_id,
                            snp["rs_id"],
                            snp["chromosome"],
                            snp["position"],
                            snp["alleles"],
                            snp["gene_name"],
                            snp["clinical_significance"],
                            ", ".join(snp["phenotypes"])
                        ])

                        print(f"RS ID: {snp['rs_id']}")
                        print(f"染色体: {snp['chromosome']}")
                        print(f"位置: {snp['position']}")
                        print(f"等位基因: {snp['alleles']}")
                        print(f"基因名称: {snp['gene_name']}")
                        print(f"临床意义: {snp['clinical_significance']}")
                        print(f"关联表型: {', '.join(snp['phenotypes'])}")
                        print("-" * 30)
                except Exception as e:
                    print(f"处理Gene ID {gene_id}的SNP信息时出错：{str(e)}")
            else:
                print(f"Gene ID {gene_id} 没有获取到SNP信息。")

            print("-" * 50)
            time.sleep(1)

        output_workbook.save(output_excel_file)
        print(f"结果已保存到: {output_excel_file}")

    except Exception as e:
        print(f"程序执行过程中出错：{str(e)}")

if __name__ == "__main__":
    main()
    print("程序执行完毕。")
    input("按回车键退出...")