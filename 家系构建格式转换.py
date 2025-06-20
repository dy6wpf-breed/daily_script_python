import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side

def clean_file_path(file_path):
    """
    清理文件路径，移除两端的双引号和多余空格
    """
    # 移除两端的空格
    file_path = file_path.strip()
    
    # 移除两端的双引号
    if file_path.startswith('"') and file_path.endswith('"'):
        file_path = file_path[1:-1]
    elif file_path.startswith("'") and file_path.endswith("'"):
        file_path = file_path[1:-1]
    
    return file_path

def reshape_to_4columns_merged(input_file, output_file):
    """
    将家系数据重新整理为4列格式，并合并相同家系的单元格
    """
    try:
        # 读取Excel文件
        print(f"正在读取文件: {os.path.basename(input_file)}")
        df = pd.read_excel(input_file, header=None)
        
        # 创建新的数据列表
        reshaped_data = []
        
        # 遍历每一行数据
        for index, row in df.iterrows():
            family_id = row[0]  # 第一列是家系编号
            
            # 收集该家系的所有个体编号
            individuals = []
            for col_idx in range(1, len(row)):
                cell_value = row[col_idx]
                if pd.notna(cell_value) and str(cell_value).strip() != '':
                    individuals.append(str(cell_value).strip())
            
            # 将个体按每4个一组分组，创建新行
            for i in range(0, len(individuals), 4):
                group = individuals[i:i+4]
                
                # 创建新行数据
                new_row = [
                    family_id,
                    group[0] if len(group) > 0 else '',
                    group[1] if len(group) > 1 else '',
                    group[2] if len(group) > 2 else '',
                    group[3] if len(group) > 3 else ''
                ]
                
                reshaped_data.append(new_row)
        
        # 使用openpyxl创建Excel文件并合并单元格
        wb = Workbook()
        ws = wb.active
        ws.title = "重构数据"
        
        # 写入数据
        for row_idx, row_data in enumerate(reshaped_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                if value != '':  # 只写入非空值
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 合并相同家系的单元格
        current_family = None
        start_row = 1
        
        for row_idx in range(1, len(reshaped_data) + 2):  # +2 to handle the last group
            if row_idx <= len(reshaped_data):
                family_id = reshaped_data[row_idx - 1][0]
            else:
                family_id = None  # 用于处理最后一组
            
            if family_id != current_family:
                # 合并前一个家系的单元格
                if current_family is not None and row_idx - 1 > start_row:
                    ws.merge_cells(f'A{start_row}:A{row_idx - 1}')
                    # 设置合并单元格的对齐方式
                    ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # 开始新的家系
                current_family = family_id
                start_row = row_idx
        
        # 设置边框和对齐
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 为所有单元格添加边框
        for row in range(1, len(reshaped_data) + 1):
            for col in range(1, 6):  # 5列
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col > 1:  # 个体编号列居中对齐
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 保存文件
        wb.save(output_file)
        
        print(f"数据转换完成！")
        print(f"原始数据: {len(df)} 行")
        print(f"转换后数据: {len(reshaped_data)} 行, 5列")
        print(f"保存到: {output_file}")
        print(f"家系单元格已合并")
        
        return len(reshaped_data)
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 {input_file}")
        print("请检查文件路径是否正确")
        return None
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        return None

def generate_output_filename(input_file):
    """
    根据输入文件名生成输出文件名
    """
    # 获取文件目录和文件名
    file_dir = os.path.dirname(input_file)
    file_name = os.path.basename(input_file)
    
    # 分离文件名和扩展名
    name_without_ext, ext = os.path.splitext(file_name)
    
    # 生成新的文件名
    output_filename = f"{name_without_ext}_4列合并{ext}"
    output_file = os.path.join(file_dir, output_filename)
    
    return output_file

# 主程序
if __name__ == "__main__":
    print("家系数据转换工具")
    print("="*60)
    print("功能: 将横向的家系数据转换为4列格式，并合并相同家系的单元格")
    print("支持格式: Excel文件 (.xlsx, .xls)")
    print("="*60)
    
    # 获取用户输入的文件路径
    print("请输入Excel文件路径:")
    print("提示: 可以直接拖拽文件到命令行窗口，或复制完整路径")
    input_file = input("文件路径: ")
    
    # 清理文件路径
    input_file = clean_file_path(input_file)
    
    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 文件不存在: {input_file}")
        # 移除重试提示，直接退出
        exit()
    
    # 检查文件扩展名 (保留警告，但不中断)
    if not input_file.lower().endswith(('.xlsx', '.xls')):
        print("警告: 文件可能不是Excel格式")
        # 移除是否继续处理的提示
    
    # 生成输出文件名
    output_file = generate_output_filename(input_file)
    
    print(f"输入文件: {input_file}")
    print(f"输出文件: {output_file}")
    
    # 移除确认是否开始处理的提示
    print("开始处理...")
    print("-" * 40)
    
    # 执行转换
    result = reshape_to_4columns_merged(input_file, output_file)
    
    if result is not None:
        print("-" * 40)
        print("✓ 转换完成！")
        print("✓ 相同家系的单元格已合并")
        print("✓ 已添加边框和居中对齐")
        print(f"✓ 输出文件: {os.path.basename(output_file)}")
    else:
        print("✗ 转换失败")
    
    # 移除询问是否继续处理其他文件的提示，程序在此结束
    print("程序结束，感谢使用！")
    input("按回车键退出...")