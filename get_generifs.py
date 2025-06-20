import sys
import os
import shutil
import re
from openpyxl import load_workbook

# Add EDirect to Python path
sys.path.insert(1, os.path.dirname(shutil.which('xtract')))
import edirect


def get_gene_generif(gene_id):
    """
    Retrieves GeneRIFs for a given Gene ID using Entrez Direct.

    Args:
        gene_id (int): NCBI Gene ID.

    Returns:
        str: GeneRIFs for the given Gene ID, separated by newlines.
            Returns an empty string if no GeneRIFs are found.
    """

    # Fetch Gene record in XML format
    xml_data = edirect.efetch(db="gene", id=gene_id, format="xml")

    # Extract GeneRIFs from XML using xtract
    generifs = edirect.pipeline(
        (f"echo '{xml_data}'",
         "xtract -pattern Gene-ref -block Gene-ref -if @source -equals 'PubMed' -element Gene-ref_text")
    )

    # Return GeneRIFs as a newline-separated string
    return generifs.replace('\t', '\n')


def main():
    """
    Reads Gene IDs from an Excel file and retrieves their GeneRIFs.
    """

    # Load the Excel workbook
    workbook = load_workbook(filename="基因注释1.xlsx")
    sheet = workbook.active

    # Iterate over rows in the Excel sheet, starting from the second row (data)
    for row in sheet.iter_rows(min_row=2):
        # Get the Gene ID from the first cell
        gene_id = row[0].value

        # Retrieve GeneRIFs for the current Gene ID
        generifs = get_gene_generif(gene_id)

        # Print the Gene ID and its GeneRIFs
        print(f"Gene ID: {gene_id}")
        print(f"GeneRIFs:\n{generifs}")
        print("-" * 20)  # Separator between gene entries


if __name__ == "__main__":
    main()