import sys
import time
from Bio import Entrez
from openpyxl import load_workbook, Workbook
import os
import xml.etree.ElementTree as ET
from googletrans import Translator

Entrez.email = "bjlmwpf@163.com"  # 请替换为您的邮箱地址


def get_gene_info(gene_id):
    try:
        print(f"正在从NCBI获取Gene ID {gene_id}的信息...")
        handle = Entrez.efetch(db="gene", id=gene_id, retmode="xml")
        xml_data = handle.read()
        handle.close()
        return xml_data
    except Exception as e:
        print(f"获取Gene ID {gene_id}的信息时出错：{str(e)}")
        return None


def extract_gene_info(xml_data):
    root = ET.fromstring(xml_data)
    result = {
        "gene_id": "",
        "gene_name": "",
        "gene_description": "",
        "generifs": []
    }

    # 提取基因ID
    gene_id_elem = root.find(".//Gene-track_geneid")
    if gene_id_elem is not None:
        result["gene_id"] = gene_id_elem.text

    # 提取基因名称
    gene_name_elem = root.find(".//Gene-ref_locus")
    if gene_name_elem is not None:
        result["gene_name"] = gene_name_elem.text

    # 提取基因描述
    gene_desc_elem = root.find(".//Gene-ref_desc")
    if gene_desc_elem is not None:
        result["gene_description"] = gene_desc_elem.text

    # 提取GeneRIF
    for comment in root.findall(".//Gene-commentary"):
        if comment.find("Gene-commentary_type").text == "18":  # GeneRIF type
            text_elem = comment.find(".//Gene-commentary_text")
            if text_elem is not None:
                result["generifs"].append(text_elem.text)

    return result


def translate_to_chinese(text):
    translator = Translator()
    try:
        return translator.translate(text, dest='zh-cn').text
    except Exception as e:
        print(f"翻译时出错：{str(e)}")
        return text


def main():
    try:
        print("程序开始执行...")

        input_excel_file = r"E:\desk\基因注释1.xlsx"
        output_excel_file = r"E:\desk\基因注释结果_GeneRIF.xlsx"
        print(f"输入Excel文件路径: {input_excel_file}")
        print(f"输出Excel文件路径: {output_excel_file}")

        if not os.path.exists(input_excel_file):
            print(f"错误：找不到文件 '{input_excel_file}'")
            return

        print(f"正在打开输入Excel文件: {input_excel_file}")
        input_workbook = load_workbook(filename=input_excel_file)
        input_sheet = input_workbook.active

        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(["Gene ID", "基因名称", "基因描述", "GeneRIF(英文)", "GeneRIF(中文)"])

        print(f"Excel文件打开成功。开始读取数据...")
        for row in input_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            gene_id = str(row[0])
            print(f"\n正在处理Gene ID: {gene_id}")

            xml_data = get_gene_info(gene_id)
            if xml_data:
                extracted_info = extract_gene_info(xml_data)

                generifs_en = "\n".join(extracted_info['generifs'])
                generifs_zh = translate_to_chinese(generifs_en) if generifs_en else ""

                output_sheet.append([
                    gene_id,
                    extracted_info['gene_name'],
                    extracted_info['gene_description'],
                    generifs_en,
                    generifs_zh
                ])

                print(f"基因名称: {extracted_info['gene_name']}")
                print(f"基因描述: {extracted_info['gene_description']}")

                if generifs_en:
                    print("GeneRIFs(英文):")
                    print(generifs_en)
                    print("GeneRIFs(中文):")
                    print(generifs_zh)
                else:
                    print("未找到GeneRIF")

            print("-" * 50)
            time.sleep(1)

        output_workbook.save(output_excel_file)
        print(f"结果已保存到: {output_excel_file}")

    except Exception as e:
        print(f"程序执行过程中出错：{str(e)}")


if __name__ == "__main__":
    main()
    print("程序执行完毕。")
    input("按回车键退出...")